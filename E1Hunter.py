from pyhunter import PyHunter
from openpyxl import Workbook
from openpyxl import load_workbook
import getpass


def Busqueda(organizacion):
    #Cantidad de resultados esperados de la búsqueda
    #El límite MENSUAL de Hunter es 25, cuidado!
    resultado = hunter.domain_search (company = organizacion, limit = 1,
                                      emails_type = 'personal')
    return resultado

print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter (apikey)
orga = input("Dominio a investigar: ")
datosEncontrados = Busqueda(orga)
if datosEncontrados is None:
    exit()
else:
    print(datosEncontrados)
    print(type(datosEncontrados))
    GuardarInformacion(datosEncontrados,orga)




def GuardarInformacion(datosEncontrados, organizacion):
    libro = Workbook()
    pagina = libro.active
    hoja = libro.create_sheet(organizacion)
    if len(libro.sheetnames)>1:
    libro.active = 1
    hoja = libro.active
else:
    hoja = libro.active
print ("La página activa es:",hoja.title)
datosimportantes = ("domain","organization","country","emails")
datosimportantes2 = ("value","type","sources","first_name","last_name","phone_number")
datosimportantes3 = ("domain","uri","extracted_on","last_seen_on")
data1 = list()
data2 = list ()
hoja["A1] = "domain"
for i in range (1,4):
     hoja[f"A{1+i}"] = datosimportantes[i]
for val in range (len(datosimportantes2[val]):
     hoja [f"A{1+val}"] = datosimportantes2[val]
for y in datosimportantes:
                  x = datosEncontrados[y]
                  data.append(x)
                  value = type(x)._name_
                  if value == "list":
                     newdata = datosEncontrados[y]
                     for data in datosimportantes2:
                     for elem in newdata:
                       i = elem[data]
                      value2 = type(i)._name_
                    data2.append(i)
                    if value2 == "list":
                     lista = elem[data]
                     for elem2 in datosimportantes3:
                     for f in lista:
                      z = f[elem2]
                     data3.append(z)
                  
data1.pop()
                  



